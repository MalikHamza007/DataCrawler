from __future__ import annotations

import csv
import json
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import sessionmaker

from app.core.config import Settings
from app.db.base import Base
from app.db.session import make_engine
from app.models import Contact, Developer, Project, SourceEvidence
from app.schemas.export import ExportCreateRequest, ExportOptions, ExportPreviewRequest
from app.services.collection_jobs import claim_next_job
from app.services.exports import create_export, export_root, generate_export_artifact, preview_export, safe_artifact_path
from app.workers.context import JobExecutionContext
from app.workers.export_handler import ExportGenerationJobHandler


def _db(tmp_path: Path):
    engine = make_engine(f"sqlite:///{tmp_path / 'exports.db'}")
    Base.metadata.create_all(bind=engine)
    SessionLocal = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    return engine, SessionLocal


def _settings(tmp_path: Path, **kwargs) -> Settings:
    return Settings(export_directory=str(tmp_path / "exports"), worker_supported_job_types=["export_generation"], **kwargs)


def _seed(session):
    developer = Developer(name="=Formula Developers", normalized_name="formula developers", classification="probable_developer", review_status="approved")
    assigned = Project(name="Gulberg Heights", normalized_name="gulberg heights", lahore_zone="Gulberg", review_status="approved", developer=developer, official_website_url="https://example.com", address="گلبرگ لاہور")
    unassigned = Project(name="Unassigned Tower", normalized_name="unassigned tower", lahore_zone="Gulberg", review_status="approved")
    session.add_all([developer, assigned, unassigned])
    session.flush()
    session.add_all([
        Contact(project_id=assigned.id, contact_type="phone", value="+923001234567", normalized_value="+923001234567", is_primary=True),
        Contact(project_id=assigned.id, contact_type="email", value="sales@example.com", is_primary=True),
        SourceEvidence(project_id=assigned.id, source_type="website", source_url="https://example.com", field_name="name", extracted_value="Gulberg Heights", captured_text="Visible public text"),
    ])
    session.commit()
    return developer, assigned, unassigned


def _request(fmt: str = "csv") -> ExportCreateRequest:
    return ExportCreateRequest(
        format=fmt,
        scope="current_project_view",
        project_filters={"lahore_zone": "Gulberg", "review_status": "approved"},
        options=ExportOptions(include_source_evidence=True),
        filename_label="Gulberg Approved Projects",
    )


def test_export_preview_has_no_side_effects(tmp_path):
    _, SessionLocal = _db(tmp_path)
    settings = _settings(tmp_path)
    with SessionLocal() as session:
        _seed(session)
        preview = preview_export(session, ExportPreviewRequest.model_validate(_request("xlsx").model_dump(exclude={"filename_label"})), settings)
        assert preview.estimated.projects == 2
        assert preview.estimated_primary_rows == 2
        assert preview.within_row_limit is True
        assert not list((tmp_path / "exports").glob("*"))


def test_create_export_queues_job_without_file(tmp_path):
    _, SessionLocal = _db(tmp_path)
    settings = _settings(tmp_path)
    with SessionLocal() as session:
        _seed(session)
        artifact = create_export(session, _request("json"), settings)
        session.commit()
        assert artifact.status == "queued"
        assert artifact.collection_job.job_type == "export_generation"
        assert artifact.filename.startswith("alduor_gulberg_approved_projects_")
        assert not safe_artifact_path(export_root(settings), artifact.internal_filename).exists()


def test_csv_export_is_one_project_per_row_and_formula_safe(tmp_path):
    _, SessionLocal = _db(tmp_path)
    settings = _settings(tmp_path)
    with SessionLocal() as session:
        _seed(session)
        artifact = create_export(session, _request("csv"), settings)
        session.commit()
        generate_export_artifact(session, artifact.id, settings)
        session.commit()
        path = safe_artifact_path(export_root(settings), artifact.internal_filename)
        with path.open("r", encoding=settings.export_csv_encoding, newline="") as handle:
            rows = list(csv.DictReader(handle))
        assert len(rows) == 2
        assert {row["Project Name"] for row in rows} == {"Gulberg Heights", "Unassigned Tower"}
        assert rows[0]["Primary Phone"].startswith("'+923001234567")
        assert artifact.sha256
        assert artifact.file_size_bytes > 0


def test_json_export_preserves_developer_project_relationships(tmp_path):
    _, SessionLocal = _db(tmp_path)
    settings = _settings(tmp_path)
    with SessionLocal() as session:
        developer, assigned, unassigned = _seed(session)
        artifact = create_export(session, _request("json"), settings)
        session.commit()
        generate_export_artifact(session, artifact.id, settings)
        session.commit()
        path = safe_artifact_path(export_root(settings), artifact.internal_filename)
        payload = json.loads(path.read_text(encoding="utf-8"))
        assert payload["schema_version"] == "m9-v1"
        dev = payload["developers"][0]
        assert dev["id"] == developer.id
        assert [project["id"] for project in dev["projects"]] == [assigned.id]
        assert [project["id"] for project in payload["unassigned_projects"]] == [unassigned.id]
        assert "گلبرگ" in payload["developers"][0]["projects"][0]["location"]["address"]


def test_xlsx_export_contains_required_sheets_and_metadata(tmp_path):
    _, SessionLocal = _db(tmp_path)
    settings = _settings(tmp_path)
    with SessionLocal() as session:
        _seed(session)
        artifact = create_export(session, _request("xlsx"), settings)
        session.commit()
        generate_export_artifact(session, artifact.id, settings)
        session.commit()
        workbook = load_workbook(safe_artifact_path(export_root(settings), artifact.internal_filename))
        assert "Dashboard" in workbook.sheetnames
        assert "Projects" in workbook.sheetnames
        assert "Developer Contacts" in workbook.sheetnames
        assert "Project Contacts" in workbook.sheetnames
        assert "Export Metadata" in workbook.sheetnames
        assert workbook["Projects"].freeze_panes == "A2"
        assert workbook["Projects"].max_row == 3
        workbook.close()


def test_export_generation_handler_completes_fixture_exports(tmp_path):
    _, SessionLocal = _db(tmp_path)
    settings = _settings(tmp_path)
    generated = {}
    for fmt in ("xlsx", "csv", "json"):
        with SessionLocal() as session:
            if not session.query(Project).count():
                _seed(session)
            artifact = create_export(session, _request(fmt), settings)
            session.commit()
            claim_next_job(session, worker_id="test-worker", settings=settings, job_id=artifact.collection_job_id)
            context = JobExecutionContext(session_factory=SessionLocal, job_id=artifact.collection_job_id, worker_id="test-worker", settings=settings)
            result = ExportGenerationJobHandler(SessionLocal, settings).execute(artifact.collection_job_id, context)
            refreshed = session.get(type(artifact), artifact.id)
            session.refresh(refreshed)
            generated[fmt] = result["file_size_bytes"]
            assert refreshed.status == "ready"
            assert refreshed.sha256
            assert safe_artifact_path(export_root(settings), refreshed.internal_filename).exists()
    assert set(generated) == {"xlsx", "csv", "json"}
