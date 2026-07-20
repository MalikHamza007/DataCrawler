from __future__ import annotations

import csv
import hashlib
import json
import os
import re
import tempfile
from dataclasses import dataclass
from datetime import UTC, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import Any
from uuid import uuid4

from fastapi import HTTPException, status
from sqlalchemy import func, or_, select
from sqlalchemy.orm import Session, selectinload

from app.core.config import Settings
from app.db.base import utc_now
from app.models import (
    CampaignEvidence,
    CollectionJob,
    CollectionLog,
    Contact,
    Developer,
    DuplicateCandidate,
    ExportArtifact,
    OutreachActivity,
    Project,
    ProjectDeveloperRelationship,
    SocialCapture,
    SocialProfile,
    SourceEvidence,
)
from app.models.intelligence import ClassificationAssessment
from app.schemas.export import (
    ExportArtifactRead,
    ExportBaseRequest,
    ExportCreateRequest,
    ExportEstimatedCounts,
    ExportOptions,
    ExportPreviewRequest,
    ExportPreviewResponse,
)
from app.services.dashboard import apply_developer_filters_for_export, apply_project_filters_for_export
from app.services.refinement import REFINEMENT_SCHEMA_VERSION, build_refined_records, refined_project_ids, refinement_summary

EXPORT_SCHEMA_VERSION = "m9-v1"
MEDIA_TYPES = {
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "csv": "text/csv; charset=utf-8",
    "json": "application/json",
}
ACTIVE_STATUSES = {"queued", "generating", "validating"}
READY_DOWNLOAD_STATUSES = {"ready"}
FORMULA_PREFIXES = ("=", "+", "-", "@")
MULTI_VALUE_SEPARATOR = " | "
PROJECT_ORDER = (Project.normalized_name.asc().nulls_last(), Project.name.asc(), Project.id.asc())
DEVELOPER_ORDER = (Developer.normalized_name.asc().nulls_last(), Developer.name.asc(), Developer.id.asc())


@dataclass
class ExportStats:
    protected_cells: int = 0
    truncated_cells: int = 0
    warnings: list[str] | None = None

    def warn(self, message: str) -> None:
        if self.warnings is None:
            self.warnings = []
        self.warnings.append(message)


def artifact_to_read(artifact: ExportArtifact) -> ExportArtifactRead:
    return ExportArtifactRead(
        id=artifact.id,
        collection_job_id=artifact.collection_job_id,
        format=artifact.format,
        scope=artifact.scope,
        status=artifact.status,
        filename=artifact.filename,
        media_type=artifact.media_type,
        row_count=artifact.row_count,
        file_size_bytes=artifact.file_size_bytes,
        sha256=artifact.sha256,
        generated_at=artifact.generated_at,
        expires_at=artifact.expires_at,
        download_count=artifact.download_count,
        last_downloaded_at=artifact.last_downloaded_at,
        error_message=artifact.error_message,
        summary=artifact.summary_json or {},
        filter_snapshot=artifact.filter_snapshot_json or {},
        options=artifact.options_json or {},
        created_at=artifact.created_at,
        updated_at=artifact.updated_at,
    )


def normalize_export_request(request: ExportBaseRequest) -> tuple[dict[str, Any], dict[str, Any], dict[str, Any]]:
    project_filters = _normalize_filters(dict(request.project_filters or {}), request.options)
    developer_filters = _normalize_filters(dict(request.developer_filters or {}), request.options)
    snapshot = {
        "project_filters": project_filters,
        "developer_filters": developer_filters,
        "requested_at": _iso(utc_now()),
    }
    return project_filters, developer_filters, snapshot


def preview_export(db: Session, request: ExportPreviewRequest, settings: Settings) -> ExportPreviewResponse:
    if not settings.export_enabled:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Export system is disabled.")
    project_filters, developer_filters, _ = normalize_export_request(request)
    counts = estimate_counts(db, request.scope, project_filters, developer_filters, request.options)
    primary = primary_row_count(request.scope, counts)
    warnings: list[str] = []
    if primary > settings.export_max_rows:
        warnings.append(f"Primary row count {primary} exceeds configured maximum {settings.export_max_rows}. Apply narrower filters.")
    if not request.options.include_source_evidence:
        warnings.append("Source evidence is omitted.")
    if request.scope == "refined_projects" and primary == 0:
        warnings.append("No complete refined records are ready. Run Prepare Clean Data and let the worker finish before exporting.")
    return ExportPreviewResponse(
        scope=request.scope,
        format=request.format,
        estimated=counts,
        estimated_primary_rows=primary,
        within_row_limit=primary <= settings.export_max_rows,
        warnings=warnings,
    )


def create_export(db: Session, request: ExportCreateRequest, settings: Settings) -> ExportArtifact:
    if not settings.export_enabled:
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Export system is disabled.")
    project_filters, developer_filters, snapshot = normalize_export_request(request)
    counts = estimate_counts(db, request.scope, project_filters, developer_filters, request.options)
    primary = primary_row_count(request.scope, counts)
    if primary > settings.export_max_rows:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"The export contains {primary} primary rows, exceeding the configured maximum of {settings.export_max_rows}. Apply Lahore zone, date or review-status filters.",
        )
    active = db.scalar(select(func.count()).select_from(ExportArtifact).where(ExportArtifact.status.in_(ACTIVE_STATUSES))) or 0
    if active >= settings.export_max_active_jobs:
        raise HTTPException(status_code=status.HTTP_429_TOO_MANY_REQUESTS, detail="The maximum number of active export jobs has been reached.")
    now = utc_now()
    job = CollectionJob(
        job_type="export_generation",
        city="Lahore",
        status="queued",
        progress_phase="queued",
        progress_message="Export queued",
        search_config_json={},
        total_items=primary,
        created_at=now,
        updated_at=now,
    )
    db.add(job)
    db.flush()
    internal_filename = f"{uuid4()}.{request.format}"
    artifact = ExportArtifact(
        collection_job_id=job.id,
        format=request.format,
        scope=request.scope,
        status="queued",
        filename=safe_export_filename(request.filename_label, request.format, now),
        internal_filename=internal_filename,
        media_type=MEDIA_TYPES[request.format],
        filter_snapshot_json=snapshot,
        options_json=request.options.model_dump(),
        summary_json={"preview": counts.model_dump(), "export_schema_version": EXPORT_SCHEMA_VERSION},
        row_count=primary,
        download_count=0,
        created_at=now,
        updated_at=now,
    )
    db.add(artifact)
    db.flush()
    job.search_config_json = {"export_artifact_id": artifact.id}
    return artifact


def retry_export(db: Session, artifact_id: int, settings: Settings) -> ExportArtifact:
    artifact = get_artifact_or_404(db, artifact_id)
    if artifact.status not in {"failed", "cancelled", "expired", "deleted"}:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Only failed, cancelled, expired or deleted exports can be retried.")
    request = ExportCreateRequest(
        format=artifact.format,
        scope=artifact.scope,
        project_filters=(artifact.filter_snapshot_json or {}).get("project_filters", {}),
        developer_filters=(artifact.filter_snapshot_json or {}).get("developer_filters", {}),
        options=ExportOptions.model_validate(artifact.options_json or {}),
        filename_label=Path(artifact.filename).stem,
    )
    return create_export(db, request, settings)


def list_exports(db: Session, *, offset: int, limit: int, filters: dict[str, Any]) -> dict[str, Any]:
    stmt = select(ExportArtifact)
    for column, key in ((ExportArtifact.format, "format"), (ExportArtifact.scope, "scope"), (ExportArtifact.status, "status")):
        if filters.get(key):
            stmt = stmt.where(column == filters[key])
    if filters.get("created_after"):
        stmt = stmt.where(ExportArtifact.created_at >= filters["created_after"])
    if filters.get("created_before"):
        stmt = stmt.where(ExportArtifact.created_at <= filters["created_before"])
    total = len(db.scalars(stmt).all())
    items = list(db.scalars(stmt.order_by(ExportArtifact.created_at.desc(), ExportArtifact.id.desc()).offset(offset).limit(min(limit, 100))).all())
    return {
        "items": [artifact_to_read(item) for item in items],
        "pagination": {
            "offset": offset,
            "limit": min(limit, 100),
            "returned": len(items),
            "total": total,
            "has_next": offset + len(items) < total,
            "has_previous": offset > 0,
        },
    }


def get_artifact_or_404(db: Session, artifact_id: int) -> ExportArtifact:
    artifact = db.get(ExportArtifact, artifact_id)
    if artifact is None:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Export artifact not found.")
    return artifact


def delete_export(db: Session, artifact: ExportArtifact, settings: Settings) -> None:
    if artifact.status in ACTIVE_STATUSES:
        raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Generating exports cannot be deleted.")
    root = export_root(settings)
    if artifact.internal_filename:
        path = safe_artifact_path(root, artifact.internal_filename)
        if path.exists():
            path.unlink()
    artifact.status = "deleted"
    artifact.internal_filename = None
    artifact.updated_at = utc_now()


def prepare_download(db: Session, artifact: ExportArtifact, settings: Settings) -> tuple[Path, str, str]:
    if artifact.status == "ready" and artifact.expires_at and _as_utc(artifact.expires_at) <= utc_now():
        artifact.status = "expired"
        db.flush()
    if artifact.status not in READY_DOWNLOAD_STATUSES:
        raise HTTPException(status_code=status.HTTP_409_CONFLICT, detail=f"Export is {artifact.status} and cannot be downloaded.")
    if not artifact.internal_filename:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Export file is unavailable.")
    path = safe_artifact_path(export_root(settings), artifact.internal_filename)
    if not path.exists() or not path.is_file():
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Export file is unavailable.")
    artifact.download_count += 1
    artifact.last_downloaded_at = utc_now()
    artifact.updated_at = utc_now()
    return path, artifact.filename, artifact.media_type


def generate_export_artifact(db: Session, artifact_id: int, settings: Settings, context: Any | None = None) -> ExportArtifact:
    artifact = get_artifact_or_404(db, artifact_id)
    if artifact.status not in {"queued", "generating"}:
        raise ValueError(f"Export artifact {artifact.id} is not queued.")
    stats = ExportStats(warnings=[])
    root = export_root(settings)
    final_path = safe_artifact_path(root, artifact.internal_filename or f"{uuid4()}.{artifact.format}")
    temp_path = final_path.with_suffix(final_path.suffix + ".part")
    artifact.status = "generating"
    artifact.updated_at = utc_now()
    db.flush()
    try:
        _progress(context, db, artifact, "planning", "Validating export options")
        data = build_export_dataset(db, artifact.scope, artifact.filter_snapshot_json or {}, ExportOptions.model_validate(artifact.options_json or {}))
        artifact.row_count = data["primary_rows"]
        _progress(context, db, artifact, "loading_data", "Loading filtered export data", total=artifact.row_count)
        temp_path.parent.mkdir(parents=True, exist_ok=True)
        if temp_path.exists():
            temp_path.unlink()
        if artifact.format == "xlsx":
            _generate_xlsx(temp_path, artifact, data, settings, stats, context=context, db=db)
        elif artifact.format == "csv":
            _generate_csv(temp_path, artifact, data, settings, stats, context=context, db=db)
        elif artifact.format == "json":
            _generate_json(temp_path, artifact, data, settings, stats, context=context, db=db)
        else:
            raise ValueError("Unsupported export format.")
        _progress(context, db, artifact, "validating", "Validating generated file")
        artifact.status = "validating"
        db.flush()
        _validate_file(temp_path, artifact, data, settings)
        _progress(context, db, artifact, "hashing", "Calculating file integrity hash")
        file_size = temp_path.stat().st_size
        if file_size > settings.export_max_file_bytes:
            raise ValueError(f"Export file exceeds configured size limit of {settings.export_max_file_bytes} bytes.")
        sha256 = _sha256(temp_path)
        _progress(context, db, artifact, "finalizing", "Finalizing download artifact")
        os.replace(temp_path, final_path)
        now = utc_now()
        summary = data["summary"]
        summary.update(
            {
                "export_schema_version": EXPORT_SCHEMA_VERSION,
                "rows_written": data["total_rows"],
                "truncated_cells": stats.truncated_cells,
                "formula_protected_cells": stats.protected_cells,
                "warnings": stats.warnings or [],
            }
        )
        artifact.status = "ready"
        artifact.file_size_bytes = file_size
        artifact.sha256 = sha256
        artifact.generated_at = now
        artifact.expires_at = now + timedelta(hours=settings.export_retention_hours)
        artifact.summary_json = summary
        artifact.updated_at = now
        _progress(context, db, artifact, "completed", "Export completed", processed=artifact.row_count, created=1)
        return artifact
    except Exception as exc:
        if settings.export_delete_temp_files and temp_path.exists():
            temp_path.unlink()
        if exc.__class__.__name__ == "JobCancellationRequested":
            artifact.status = "cancelled"
            artifact.error_message = None
        else:
            artifact.status = "failed"
            artifact.error_message = str(exc)[:2000]
        artifact.updated_at = utc_now()
        db.flush()
        raise


def build_export_dataset(db: Session, scope: str, filter_snapshot: dict[str, Any], options: ExportOptions) -> dict[str, Any]:
    project_filters = dict((filter_snapshot or {}).get("project_filters") or {})
    developer_filters = dict((filter_snapshot or {}).get("developer_filters") or {})
    project_ids = _project_ids_for_scope(db, scope, project_filters, options)
    developer_ids = _developer_ids_for_scope(db, scope, developer_filters, project_ids, options)
    projects = _load_projects(db, project_ids)
    developers = _load_developers(db, developer_ids)
    contacts = _load_contacts(db, developer_ids, project_ids) if options.include_contacts or scope == "refined_projects" else []
    social_profiles = _load_social_profiles(db, developer_ids, project_ids) if options.include_social_profiles else []
    campaigns = _load_campaigns(db, developer_ids, project_ids) if options.include_campaign_evidence else []
    relationships = _load_relationships(db, developer_ids, project_ids) if options.include_relationships else []
    duplicates = _load_duplicates(db, developer_ids, project_ids) if options.include_duplicate_candidates else []
    outreach = _load_outreach(db, developer_ids, project_ids) if options.include_outreach_activities else []
    evidence = _load_evidence(db, developer_ids, project_ids) if options.include_source_evidence else []
    social_captures = _load_social_captures(db, developer_ids, project_ids) if options.include_unassigned_social_captures else []
    collection_logs = _load_collection_logs(db) if options.include_collection_logs else []
    review_queue = _review_queue(projects, developers, relationships, duplicates, social_captures)
    refined_records = build_refined_records(db, project_ids) if scope == "refined_projects" else []
    summary = {
        "projects_exported": len(projects),
        "developers_exported": len(developers),
        "contacts_exported": len(contacts),
        "social_profiles_exported": len(social_profiles),
        "campaign_evidence_exported": len(campaigns),
        "source_evidence_exported": len(evidence),
        "relationships_exported": len(relationships),
        "duplicate_candidates_exported": len(duplicates),
        "outreach_activities_exported": len(outreach),
        "social_captures_exported": len(social_captures),
        "collection_logs_exported": len(collection_logs),
    }
    if scope == "refined_projects":
        summary.update(refinement_summary(db))
        summary["refined_projects_exported"] = len(refined_records)
    primary = _primary_from_scope(scope, projects, developers, review_queue, outreach)
    return {
        "projects": projects,
        "developers": developers,
        "contacts": contacts,
        "social_profiles": social_profiles,
        "campaigns": campaigns,
        "relationships": relationships,
        "duplicates": duplicates,
        "outreach": outreach,
        "evidence": evidence,
        "social_captures": social_captures,
        "collection_logs": collection_logs,
        "review_queue": review_queue,
        "refined_records": refined_records,
        "project_ids": project_ids,
        "developer_ids": developer_ids,
        "primary_rows": primary,
        "summary": summary,
        "total_rows": sum(value for value in summary.values() if isinstance(value, int)),
    }


def estimate_counts(db: Session, scope: str, project_filters: dict[str, Any], developer_filters: dict[str, Any], options: ExportOptions) -> ExportEstimatedCounts:
    project_ids = _project_ids_for_scope(db, scope, project_filters, options)
    developer_ids = _developer_ids_for_scope(db, scope, developer_filters, project_ids, options)
    return ExportEstimatedCounts(
        projects=len(project_ids),
        developers=len(developer_ids),
        project_contacts=_count_where(db, Contact, Contact.project_id, project_ids) if options.include_contacts else 0,
        developer_contacts=_count_where(db, Contact, Contact.developer_id, developer_ids) if options.include_contacts else 0,
        social_profiles=(_count_where(db, SocialProfile, SocialProfile.project_id, project_ids) + _count_where(db, SocialProfile, SocialProfile.developer_id, developer_ids)) if options.include_social_profiles else 0,
        campaign_evidence=(_count_where(db, CampaignEvidence, CampaignEvidence.project_id, project_ids) + _count_where(db, CampaignEvidence, CampaignEvidence.developer_id, developer_ids)) if options.include_campaign_evidence else 0,
        source_evidence=(_count_where(db, SourceEvidence, SourceEvidence.project_id, project_ids) + _count_where(db, SourceEvidence, SourceEvidence.developer_id, developer_ids)) if options.include_source_evidence else 0,
        relationships=len(_load_relationships(db, developer_ids, project_ids)) if options.include_relationships else 0,
        duplicate_candidates=len(_load_duplicates(db, developer_ids, project_ids)) if options.include_duplicate_candidates else 0,
        outreach_activities=(_count_where(db, OutreachActivity, OutreachActivity.project_id, project_ids) + _count_where(db, OutreachActivity, OutreachActivity.developer_id, developer_ids)) if options.include_outreach_activities else 0,
        social_captures=(_count_where(db, SocialCapture, SocialCapture.project_id, project_ids) + _count_where(db, SocialCapture, SocialCapture.developer_id, developer_ids)) if options.include_unassigned_social_captures else 0,
        collection_logs=(db.scalar(select(func.count()).select_from(CollectionLog)) or 0) if options.include_collection_logs else 0,
    )


def primary_row_count(scope: str, counts: ExportEstimatedCounts) -> int:
    if scope in {"current_developer_view", "all_developers"}:
        return counts.developers
    if scope == "full_intelligence":
        return counts.projects + counts.developers
    if scope == "outreach_pipeline":
        return counts.projects + counts.developers
    if scope == "review_queue":
        return counts.projects + counts.developers + counts.relationships + counts.duplicate_candidates + counts.social_captures
    return counts.projects


def export_root(settings: Settings) -> Path:
    root = Path(settings.export_directory)
    if not root.is_absolute():
        root = Path.cwd() / root
    resolved = root.resolve()
    resolved.mkdir(parents=True, exist_ok=True)
    return resolved


def safe_artifact_path(root: Path, internal_filename: str) -> Path:
    if not internal_filename or "/" in internal_filename or "\\" in internal_filename or "\x00" in internal_filename:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid export filename.")
    path = (root / internal_filename).resolve()
    if root not in path.parents and path != root:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="Invalid export path.")
    return path


def safe_export_filename(label: str | None, fmt: str, now: datetime) -> str:
    base = (label or "lahore_export").strip().lower()
    base = re.sub(r"\.[A-Za-z0-9]{1,8}$", "", base)
    base = re.sub(r"[^a-z0-9\s_-]+", "", base)
    base = re.sub(r"[\s_-]+", "_", base).strip("_")[:60]
    if not base:
        base = "lahore_export"
    stamp = _as_utc(now).strftime("%Y%m%dT%H%M%SZ")
    return f"alduor_{base}_{stamp}.{fmt}"


def spreadsheet_text(value: Any, stats: ExportStats, settings: Settings, *, untrusted: bool = True) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float, bool, datetime)):
        return value
    if isinstance(value, Decimal):
        return float(value)
    text = clean_text(str(value))
    limit = settings.export_excel_max_cell_characters
    if len(text) > limit:
        suffix = "\n[Truncated for export]"
        text = text[: max(0, limit - len(suffix))] + suffix
        stats.truncated_cells += 1
    if settings.export_formula_protection and untrusted and _is_formula_like(text):
        text = "'" + text
        stats.protected_cells += 1
    return text


def clean_text(value: str) -> str:
    value = value.replace("\x00", "")
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    return re.sub(r"[\x01-\x08\x0b\x0c\x0e-\x1f]", "", value).strip()


def cleanup_expired_exports(db: Session, settings: Settings, *, dry_run: bool = False) -> dict[str, int]:
    root = export_root(settings)
    now = utc_now()
    expired = list(db.scalars(select(ExportArtifact).where(ExportArtifact.status == "ready", ExportArtifact.expires_at <= now)).all())
    files_removed = 0
    artifacts_expired = 0
    for artifact in expired:
        if artifact.internal_filename:
            path = safe_artifact_path(root, artifact.internal_filename)
            if path.exists() and not dry_run:
                path.unlink()
                files_removed += 1
            elif path.exists():
                files_removed += 1
        if not dry_run:
            artifact.status = "expired"
            artifact.updated_at = now
        artifacts_expired += 1
    orphan_parts = 0
    cutoff = now.timestamp() - 3600
    for path in root.glob("*.part"):
        if path.resolve().parent != root:
            continue
        if path.stat().st_mtime < cutoff:
            orphan_parts += 1
            if not dry_run:
                path.unlink()
    return {"artifacts_expired": artifacts_expired, "files_removed": files_removed, "orphan_part_files_removed": orphan_parts if not dry_run else 0, "orphan_part_files_found": orphan_parts}


def _generate_xlsx(path: Path, artifact: ExportArtifact, data: dict[str, Any], settings: Settings, stats: ExportStats, *, context: Any | None, db: Session) -> None:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.table import Table, TableStyleInfo

    _progress(context, db, artifact, "generating_workbook", "Generating Excel workbook")
    wb = Workbook()
    wb.remove(wb.active)
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    def write_sheet(name: str, headers: list[str], rows: list[dict[str, Any]], table_name: str | None = None) -> None:
        _check_cancelled(context)
        ws = wb.create_sheet(name)
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        for row in rows:
            ws.append([spreadsheet_text(row.get(header), stats, settings) for header in headers])
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for column_cells in ws.columns:
            letter = column_cells[0].column_letter
            width = min(max(len(str(column_cells[0].value or "")) + 4, 12), 60)
            ws.column_dimensions[letter].width = width
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                if isinstance(cell.value, str) and cell.value.startswith("http"):
                    cell.hyperlink = cell.value
                    cell.style = "Hyperlink"
        if table_name and headers:
            table = Table(displayName=table_name, ref=ws.dimensions)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
            ws.add_table(table)

    write_sheet("Dashboard", ["Metric", "Value"], _dashboard_rows(artifact, data, stats), None)
    if artifact.scope in {"full_intelligence", "current_developer_view", "all_developers", "outreach_pipeline"}:
        write_sheet("Developers", DEVELOPER_HEADERS, _developer_rows(data), "tblDevelopers")
    if artifact.scope == "refined_projects":
        write_sheet("Refined Projects", REFINED_HEADERS, _refined_csv_rows(data), "tblRefinedProjects")
    elif artifact.scope in {"full_intelligence", "current_project_view", "all_projects", "review_queue", "outreach_pipeline"}:
        write_sheet("Projects", PROJECT_HEADERS, _project_rows(data), "tblProjects")
    if data["contacts"]:
        write_sheet("Developer Contacts", DEVELOPER_CONTACT_HEADERS, _developer_contact_rows(data), "tblDeveloperContacts")
        write_sheet("Project Contacts", PROJECT_CONTACT_HEADERS, _project_contact_rows(data), "tblProjectContacts")
    if data["social_profiles"]:
        write_sheet("Social Profiles", SOCIAL_HEADERS, _social_rows(data), "tblSocialProfiles")
    if data["campaigns"]:
        write_sheet("Campaign Evidence", CAMPAIGN_HEADERS, _campaign_rows(data), "tblCampaignEvidence")
    if data["evidence"]:
        write_sheet("Source Evidence", EVIDENCE_HEADERS, _evidence_rows(data), "tblSourceEvidence")
    if data["relationships"]:
        write_sheet("Relationships", RELATIONSHIP_HEADERS, _relationship_rows(data), "tblRelationships")
    if data["review_queue"]:
        write_sheet("Review Queue", REVIEW_HEADERS, data["review_queue"], "tblReviewQueue")
    excluded = _excluded_rows(data)
    if excluded:
        write_sheet("Excluded Brokers", EXCLUDED_HEADERS, excluded, "tblExcludedBrokers")
    if data["duplicates"]:
        write_sheet("Duplicates", DUPLICATE_HEADERS, _duplicate_rows(data), "tblDuplicateCandidates")
    if data["outreach"]:
        write_sheet("Outreach", OUTREACH_HEADERS, _outreach_rows(data), "tblOutreach")
    if data["collection_logs"]:
        write_sheet("Collection Logs", LOG_HEADERS, _collection_log_rows(data), "tblCollectionLogs")
    write_sheet("Export Metadata", ["Key", "Value"], _metadata_rows(artifact, data, stats), None)
    wb.save(path)
    if settings.export_validate_files:
        with path.open("rb") as handle:
            load_workbook(handle).close()


def _generate_csv(path: Path, artifact: ExportArtifact, data: dict[str, Any], settings: Settings, stats: ExportStats, *, context: Any | None, db: Session) -> None:
    _progress(context, db, artifact, "generating_csv", "Generating CSV export")
    headers = REFINED_HEADERS if artifact.scope == "refined_projects" else CSV_HEADERS
    rows = _refined_csv_rows(data) if artifact.scope == "refined_projects" else _csv_project_rows(data)
    with path.open("w", encoding=settings.export_csv_encoding, newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        for row in rows:
            _check_cancelled(context)
            writer.writerow({header: spreadsheet_text(row.get(header), stats, settings) for header in headers})


def _generate_json(path: Path, artifact: ExportArtifact, data: dict[str, Any], settings: Settings, stats: ExportStats, *, context: Any | None, db: Session) -> None:
    _progress(context, db, artifact, "generating_json", "Generating JSON export")
    payload = _refined_json_payload(artifact, data) if artifact.scope == "refined_projects" else _json_payload(artifact, data)
    _check_cancelled(context)
    with path.open("w", encoding="utf-8") as handle:
        json.dump(payload, handle, ensure_ascii=False, indent=settings.export_json_indent)
        handle.write("\n")


def _validate_file(path: Path, artifact: ExportArtifact, data: dict[str, Any], settings: Settings) -> None:
    if not settings.export_validate_files:
        return
    if artifact.format == "xlsx":
        from openpyxl import load_workbook

        with path.open("rb") as handle:
            wb = load_workbook(handle, read_only=False)
        if "Dashboard" not in wb.sheetnames or "Export Metadata" not in wb.sheetnames:
            raise ValueError("Excel validation failed: required sheets missing.")
        if len(wb.sheetnames) != len(set(wb.sheetnames)):
            raise ValueError("Excel validation failed: duplicate sheet names.")
        if "Projects" in wb.sheetnames and wb["Projects"].max_row - 1 != len(data["projects"]):
            raise ValueError("Excel validation failed: project row count mismatch.")
        for name in wb.sheetnames:
            if wb[name].max_row >= 1 and not wb[name].freeze_panes and name not in {"Dashboard", "Export Metadata"}:
                raise ValueError("Excel validation failed: frozen panes missing.")
        wb.close()
    elif artifact.format == "csv":
        with path.open("r", encoding=settings.export_csv_encoding, newline="") as handle:
            reader = csv.DictReader(handle)
            rows = list(reader)
        expected_headers = REFINED_HEADERS if artifact.scope == "refined_projects" else CSV_HEADERS
        if reader.fieldnames != expected_headers:
            raise ValueError("CSV validation failed: headers mismatch.")
        if len(rows) != len(data["projects"]):
            raise ValueError("CSV validation failed: project row count mismatch.")
        if rows and not rows[0].get("Project ID"):
            raise ValueError("CSV validation failed: missing project ID.")
    elif artifact.format == "json":
        with path.open("r", encoding="utf-8") as handle:
            payload = json.load(handle)
        expected_schema = REFINEMENT_SCHEMA_VERSION if artifact.scope == "refined_projects" else EXPORT_SCHEMA_VERSION
        if payload.get("schema_version") != expected_schema:
            raise ValueError("JSON validation failed: schema version missing.")
        if artifact.scope == "refined_projects":
            if len(payload.get("projects", [])) != len(data["projects"]):
                raise ValueError("JSON validation failed: refined project count mismatch.")
            for project in payload.get("projects", []):
                if not all((project.get("project_name"), project.get("developer", {}).get("name"), project.get("developer", {}).get("phone"), project.get("developer", {}).get("email"), project.get("location", {}).get("address"))):
                    raise ValueError("JSON validation failed: refined project is incomplete.")
            return
        project_ids: set[int] = set()
        for developer in payload.get("developers", []):
            for project in developer.get("projects", []):
                if project["id"] in project_ids:
                    raise ValueError("JSON validation failed: duplicate project.")
                project_ids.add(project["id"])
        for project in payload.get("unassigned_projects", []):
            if project["id"] in project_ids:
                raise ValueError("JSON validation failed: unassigned project duplicated.")
            project_ids.add(project["id"])
        if len(project_ids) != len(data["projects"]):
            raise ValueError("JSON validation failed: project count mismatch.")


def _normalize_filters(filters: dict[str, Any], options: ExportOptions) -> dict[str, Any]:
    normalized = {key: _coerce_filter_value(value) for key, value in filters.items() if value not in (None, "")}
    if options.include_merged_records:
        normalized["include_merged"] = True
    if "include_merged" in normalized:
        normalized["include_merged"] = _truthy(normalized["include_merged"])
    return normalized


def _coerce_filter_value(value: Any) -> Any:
    if isinstance(value, str):
        lowered = value.lower()
        if lowered == "true":
            return True
        if lowered == "false":
            return False
    return value


def _project_ids_for_scope(db: Session, scope: str, filters: dict[str, Any], options: ExportOptions) -> list[int]:
    if scope in {"current_developer_view", "all_developers"}:
        return []
    scoped = dict(filters)
    if scope == "refined_projects":
        return refined_project_ids(db)
    if scope == "all_projects":
        scoped = {}
    if scope == "review_queue":
        scoped["review_status"] = scoped.get("review_status") or "needs_review"
    if scope == "outreach_pipeline":
        scoped["review_status"] = scoped.get("review_status") or "approved"
    stmt = apply_project_filters_for_export(select(Project.id, Project.normalized_name, Project.name), scoped)
    rows = db.execute(stmt.order_by(*PROJECT_ORDER)).all()
    ids = [row[0] for row in rows]
    if not options.include_rejected_records:
        ids = [pid for pid in ids if (db.get(Project, pid).review_status != "rejected")]
    if not options.include_excluded_records:
        ids = [pid for pid in ids if (db.get(Project, pid).review_status != "excluded")]
    return ids


def _developer_ids_for_scope(db: Session, scope: str, filters: dict[str, Any], project_ids: list[int], options: ExportOptions) -> list[int]:
    if scope in {"current_project_view", "all_projects", "refined_projects"}:
        ids = [row[0] for row in db.execute(select(Project.developer_id).where(Project.id.in_(project_ids), Project.developer_id.is_not(None)).distinct()).all()] if project_ids else []
        return sorted(ids)
    scoped = dict(filters)
    if scope == "all_developers":
        scoped = {}
    if scope == "review_queue":
        scoped["review_status"] = scoped.get("review_status") or "needs_review"
    if scope == "outreach_pipeline":
        scoped["review_status"] = scoped.get("review_status") or "approved"
    stmt = apply_developer_filters_for_export(select(Developer.id, Developer.normalized_name, Developer.name), scoped)
    ids = [row[0] for row in db.execute(stmt.order_by(*DEVELOPER_ORDER)).all()]
    if scope == "full_intelligence":
        related = [row[0] for row in db.execute(select(Project.developer_id).where(Project.id.in_(project_ids), Project.developer_id.is_not(None)).distinct()).all()] if project_ids else []
        ids = sorted(set(ids) | set(related))
    if not options.include_rejected_records:
        ids = [did for did in ids if (db.get(Developer, did).review_status != "rejected")]
    if not options.include_excluded_records:
        ids = [did for did in ids if (db.get(Developer, did).review_status != "excluded")]
    return ids


def _load_projects(db: Session, ids: list[int]) -> list[Project]:
    if not ids:
        return []
    projects = list(db.scalars(select(Project).options(selectinload(Project.developer)).where(Project.id.in_(ids))).all())
    order = {pid: index for index, pid in enumerate(ids)}
    return sorted(projects, key=lambda item: order[item.id])


def _load_developers(db: Session, ids: list[int]) -> list[Developer]:
    if not ids:
        return []
    developers = list(db.scalars(select(Developer).where(Developer.id.in_(ids))).all())
    order = {did: index for index, did in enumerate(ids)}
    return sorted(developers, key=lambda item: order[item.id])


def _load_contacts(db: Session, developer_ids: list[int], project_ids: list[int]) -> list[Contact]:
    return list(db.scalars(select(Contact).where(or_(Contact.developer_id.in_(developer_ids or [-1]), Contact.project_id.in_(project_ids or [-1]))).order_by(Contact.developer_id, Contact.project_id, Contact.contact_type, Contact.id)).all())


def _load_social_profiles(db: Session, developer_ids: list[int], project_ids: list[int]) -> list[SocialProfile]:
    return list(db.scalars(select(SocialProfile).where(or_(SocialProfile.developer_id.in_(developer_ids or [-1]), SocialProfile.project_id.in_(project_ids or [-1]))).order_by(SocialProfile.developer_id, SocialProfile.project_id, SocialProfile.platform, SocialProfile.id)).all())


def _load_campaigns(db: Session, developer_ids: list[int], project_ids: list[int]) -> list[CampaignEvidence]:
    return list(db.scalars(select(CampaignEvidence).where(or_(CampaignEvidence.developer_id.in_(developer_ids or [-1]), CampaignEvidence.project_id.in_(project_ids or [-1]))).order_by(CampaignEvidence.first_seen_at.desc(), CampaignEvidence.id)).all())


def _load_relationships(db: Session, developer_ids: list[int], project_ids: list[int]) -> list[ProjectDeveloperRelationship]:
    return list(db.scalars(select(ProjectDeveloperRelationship).where(or_(ProjectDeveloperRelationship.developer_id.in_(developer_ids or [-1]), ProjectDeveloperRelationship.project_id.in_(project_ids or [-1]))).order_by(ProjectDeveloperRelationship.project_id, ProjectDeveloperRelationship.system_score.desc(), ProjectDeveloperRelationship.id)).all())


def _load_duplicates(db: Session, developer_ids: list[int], project_ids: list[int]) -> list[DuplicateCandidate]:
    return list(db.scalars(select(DuplicateCandidate).where(or_(DuplicateCandidate.left_developer_id.in_(developer_ids or [-1]), DuplicateCandidate.right_developer_id.in_(developer_ids or [-1]), DuplicateCandidate.left_project_id.in_(project_ids or [-1]), DuplicateCandidate.right_project_id.in_(project_ids or [-1]))).order_by(DuplicateCandidate.duplicate_score.desc(), DuplicateCandidate.id)).all())


def _load_outreach(db: Session, developer_ids: list[int], project_ids: list[int]) -> list[OutreachActivity]:
    return list(db.scalars(select(OutreachActivity).where(or_(OutreachActivity.developer_id.in_(developer_ids or [-1]), OutreachActivity.project_id.in_(project_ids or [-1]))).order_by(OutreachActivity.occurred_at.desc(), OutreachActivity.id)).all())


def _load_evidence(db: Session, developer_ids: list[int], project_ids: list[int]) -> list[SourceEvidence]:
    return list(db.scalars(select(SourceEvidence).where(or_(SourceEvidence.developer_id.in_(developer_ids or [-1]), SourceEvidence.project_id.in_(project_ids or [-1]))).order_by(SourceEvidence.collected_at.desc(), SourceEvidence.id)).all())


def _load_social_captures(db: Session, developer_ids: list[int], project_ids: list[int]) -> list[SocialCapture]:
    return list(db.scalars(select(SocialCapture).where(or_(SocialCapture.developer_id.in_(developer_ids or [-1]), SocialCapture.project_id.in_(project_ids or [-1]), SocialCapture.review_status == "unassigned")).order_by(SocialCapture.captured_at.desc(), SocialCapture.id)).all())


def _load_collection_logs(db: Session) -> list[CollectionLog]:
    return list(db.scalars(select(CollectionLog).order_by(CollectionLog.created_at.desc(), CollectionLog.id.desc()).limit(5000)).all())


def _count_where(db: Session, model: Any, column: Any, ids: list[int]) -> int:
    if not ids:
        return 0
    return db.scalar(select(func.count()).select_from(model).where(column.in_(ids))) or 0


def _primary_from_scope(scope: str, projects: list[Project], developers: list[Developer], review_queue: list[dict[str, Any]], outreach: list[OutreachActivity]) -> int:
    if scope in {"current_developer_view", "all_developers"}:
        return len(developers)
    if scope == "full_intelligence":
        return len(projects) + len(developers)
    if scope == "review_queue":
        return len(review_queue)
    if scope == "outreach_pipeline":
        return len(projects) + len(developers)
    return len(projects)


def _classifications_by_owner(db_objects: list[Any], entity_type: str) -> dict[int, ClassificationAssessment]:
    # Loaded lazily per row below would be noisy, so this placeholder keeps rows simple when no assessment is preloaded.
    return {}


def _latest_assessment_for_project(project_id: int, data: dict[str, Any]) -> ClassificationAssessment | None:
    return None


def _latest_assessment_for_developer(developer_id: int, data: dict[str, Any]) -> ClassificationAssessment | None:
    return None


def _project_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    contacts = _group(data["contacts"], "project_id")
    profiles = _group(data["social_profiles"], "project_id")
    campaigns = _group(data["campaigns"], "project_id")
    evidence = _group(data["evidence"], "project_id")
    rels = _group(data["relationships"], "project_id")
    dups = _project_duplicate_map(data["duplicates"])
    rows = []
    for project in data["projects"]:
        primary_phone = _primary_contact(contacts.get(project.id, []), ["mobile", "phone", "landline"])
        primary_whatsapp = _primary_contact(contacts.get(project.id, []), ["whatsapp"])
        primary_email = _primary_contact(contacts.get(project.id, []), ["email"])
        row = {
            "Project ID": project.id,
            "Project Name": project.name,
            "Developer ID": project.developer_id,
            "Developer Name": project.developer.name if project.developer else "",
            "Candidate Developer Count": len(rels.get(project.id, [])),
            "Project Type": project.project_type,
            "Project Status": project.project_status,
            "Effective Classification": project.project_type,
            "System Suggested Classification": "",
            "Classification Score": "",
            "Classification Confidence": "",
            "Verification Status": project.verification_status,
            "Review Status": project.review_status,
            "Record Status": project.record_status,
            "Lahore Zone": project.lahore_zone,
            "Location Scope": project.city,
            "Address": project.address,
            "Latitude": project.latitude,
            "Longitude": project.longitude,
            "Google Place ID": project.google_place_id,
            "Google Maps URL": project.google_maps_url,
            "Official Website": project.official_website_url,
            "Primary Phone": primary_phone.value if primary_phone else "",
            "Primary WhatsApp": primary_whatsapp.value if primary_whatsapp else "",
            "Primary Email": primary_email.value if primary_email else "",
            "Facebook": _profile_url(profiles.get(project.id, []), "facebook"),
            "Instagram": _profile_url(profiles.get(project.id, []), "instagram"),
            "LinkedIn": _profile_url(profiles.get(project.id, []), "linkedin"),
            "X": _profile_url(profiles.get(project.id, []), "x"),
            "YouTube": _profile_url(profiles.get(project.id, []), "youtube"),
            "TikTok": _profile_url(profiles.get(project.id, []), "tiktok"),
            "Campaign Evidence Count": len(campaigns.get(project.id, [])),
            "Source Evidence Count": len(evidence.get(project.id, [])),
            "Pending Relationship": any(rel.status == "candidate" for rel in rels.get(project.id, [])),
            "Pending Duplicate": project.id in dups,
            "Outreach Status": project.outreach_status,
            "Last Outreach": _iso(project.last_outreach_at),
            "Next Follow-up": _iso(project.next_follow_up_at),
            "Created At": _iso(project.created_at),
            "Updated At": _iso(project.updated_at),
            "Merged Into Project ID": project.merged_into_project_id,
        }
        rows.append(row)
    return rows


def _developer_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    contacts = _group(data["contacts"], "developer_id")
    profiles = _group(data["social_profiles"], "developer_id")
    projects_by_developer = _group(data["projects"], "developer_id")
    evidence = _group(data["evidence"], "developer_id")
    dups = _developer_duplicate_map(data["duplicates"])
    rows = []
    for developer in data["developers"]:
        primary_phone = _primary_contact(contacts.get(developer.id, []), ["mobile", "phone", "landline"])
        primary_whatsapp = _primary_contact(contacts.get(developer.id, []), ["whatsapp"])
        primary_email = _primary_contact(contacts.get(developer.id, []), ["email"])
        dev_projects = projects_by_developer.get(developer.id, [])
        rows.append({
            "Developer ID": developer.id,
            "Developer Name": developer.name,
            "Legal Name": developer.legal_name,
            "Effective Classification": developer.classification,
            "System Suggested Classification": "",
            "Classification Score": "",
            "Classification Confidence": "",
            "Classification Review Status": "",
            "Verification Status": developer.verification_status,
            "Review Status": developer.review_status,
            "Record Status": developer.record_status,
            "Website": developer.website_url,
            "Office Address": developer.office_address,
            "City": developer.city,
            "Country": developer.country,
            "Primary Phone": primary_phone.value if primary_phone else "",
            "Primary WhatsApp": primary_whatsapp.value if primary_whatsapp else "",
            "Primary Email": primary_email.value if primary_email else "",
            "Facebook": _profile_url(profiles.get(developer.id, []), "facebook"),
            "Instagram": _profile_url(profiles.get(developer.id, []), "instagram"),
            "LinkedIn": _profile_url(profiles.get(developer.id, []), "linkedin"),
            "X": _profile_url(profiles.get(developer.id, []), "x"),
            "YouTube": _profile_url(profiles.get(developer.id, []), "youtube"),
            "TikTok": _profile_url(profiles.get(developer.id, []), "tiktok"),
            "Total Projects": len(dev_projects),
            "Lahore Projects": len([project for project in dev_projects if project.city == "Lahore"]),
            "Approved Projects": len([project for project in dev_projects if project.review_status == "approved"]),
            "Pending Relationships": "",
            "Pending Duplicates": developer.id in dups,
            "Evidence Count": len(evidence.get(developer.id, [])),
            "Outreach Status": developer.outreach_status,
            "Last Outreach": _iso(developer.last_outreach_at),
            "Next Follow-up": _iso(developer.next_follow_up_at),
            "Created At": _iso(developer.created_at),
            "Updated At": _iso(developer.updated_at),
            "Merged Into Developer ID": developer.merged_into_developer_id,
        })
    return rows


PROJECT_HEADERS = ["Project ID", "Project Name", "Developer ID", "Developer Name", "Candidate Developer Count", "Project Type", "Project Status", "Effective Classification", "System Suggested Classification", "Classification Score", "Classification Confidence", "Verification Status", "Review Status", "Record Status", "Lahore Zone", "Location Scope", "Address", "Latitude", "Longitude", "Google Place ID", "Google Maps URL", "Official Website", "Primary Phone", "Primary WhatsApp", "Primary Email", "Facebook", "Instagram", "LinkedIn", "X", "YouTube", "TikTok", "Campaign Evidence Count", "Source Evidence Count", "Pending Relationship", "Pending Duplicate", "Outreach Status", "Last Outreach", "Next Follow-up", "Created At", "Updated At", "Merged Into Project ID"]
DEVELOPER_HEADERS = ["Developer ID", "Developer Name", "Legal Name", "Effective Classification", "System Suggested Classification", "Classification Score", "Classification Confidence", "Classification Review Status", "Verification Status", "Review Status", "Record Status", "Website", "Office Address", "City", "Country", "Primary Phone", "Primary WhatsApp", "Primary Email", "Facebook", "Instagram", "LinkedIn", "X", "YouTube", "TikTok", "Total Projects", "Lahore Projects", "Approved Projects", "Pending Relationships", "Pending Duplicates", "Evidence Count", "Outreach Status", "Last Outreach", "Next Follow-up", "Created At", "Updated At", "Merged Into Developer ID"]
CSV_HEADERS = ["Project ID", "Project Name", "Developer ID", "Developer Name", "Developer Classification", "Project Classification", "Classification Score", "Review Status", "Verification Status", "Project Type", "Project Status", "Lahore Zone", "Address", "Latitude", "Longitude", "Google Place ID", "Google Maps URL", "Official Website", "Primary Phone", "Other Phones", "Primary WhatsApp", "Other WhatsApp", "Primary Email", "Other Emails", "Facebook", "Instagram", "LinkedIn", "X", "YouTube", "TikTok", "Campaign Evidence Count", "Latest Campaign Platform", "Latest Campaign CTA", "Latest Campaign Source", "Pending Relationship", "Pending Duplicate", "Evidence Count", "Outreach Status", "Last Outreach", "Next Follow-up", "Created At", "Updated At"]
REFINED_HEADERS = ["Project ID", "Project Name", "Developer ID", "Developer Name", "Developer Phone", "Developer Email", "Developer Website", "Project Phone", "Project Email", "Project Website", "Exact Address", "Latitude", "Longitude", "Google Place ID", "Google Maps URL", "Refinement Score"]
DEVELOPER_CONTACT_HEADERS = ["Contact ID", "Developer ID", "Developer Name", "Contact Type", "Label", "Original Value", "Normalized Value", "Contact Person", "Designation", "Primary", "Public Business Contact", "Verification Status", "Source URL", "Created At", "Updated At"]
PROJECT_CONTACT_HEADERS = ["Contact ID", "Project ID", "Project Name", "Developer Name", "Contact Type", "Label", "Original Value", "Normalized Value", "Contact Person", "Designation", "Primary", "Public Business Contact", "Verification Status", "Source URL", "Created At", "Updated At"]
SOCIAL_HEADERS = ["Social Profile ID", "Owner Type", "Owner ID", "Owner Name", "Platform", "Profile Name", "Profile URL", "Normalized URL", "Official", "Verification Status", "Source Evidence Count", "Created At", "Updated At"]
CAMPAIGN_HEADERS = ["Campaign Evidence ID", "Developer ID", "Developer Name", "Project ID", "Project Name", "Platform", "Campaign Type", "Advertiser Name", "Campaign Text", "Call to Action", "Destination URL", "Visible Status", "Verification Status", "First Seen At", "Last Seen At", "Source URL", "Social Capture ID", "Created At", "Updated At"]
EVIDENCE_HEADERS = ["Evidence ID", "Developer ID", "Developer Name", "Project ID", "Project Name", "Collection Job ID", "Source Type", "Source URL", "Source Title", "Field Name", "Extracted Value", "Relevant Excerpt", "Verification Status", "Collected At", "Created At"]
RELATIONSHIP_HEADERS = ["Relationship ID", "Project ID", "Project Name", "Developer ID", "Developer Name", "Relationship Type", "Status", "System Score", "Confidence", "Rule Version", "Explanation", "Source URL", "Evidence Count", "Evaluated At", "Reviewed At", "Review Note"]
REVIEW_HEADERS = ["Entity Type", "Entity ID", "Entity Name", "Review Status", "Effective Classification", "Suggested Classification", "Classification Score", "Pending Relationship", "Pending Duplicate", "Missing Developer", "Missing Phone", "Missing WhatsApp", "Missing Email", "Missing Website", "Missing Coordinates", "Conflicting Evidence", "Latest Warning", "Last Reviewed At", "Created At", "Updated At"]
EXCLUDED_HEADERS = ["Entity Type", "Entity ID", "Entity Name", "Classification", "Review Status", "Exclusion or Rejection Reason", "Classification Score", "Top Positive Signal", "Top Negative Signal", "Website", "Phone", "Lahore Zone", "Evidence Count", "Reviewed At", "Updated At"]
DUPLICATE_HEADERS = ["Duplicate Candidate ID", "Entity Type", "Left ID", "Left Name", "Right ID", "Right Name", "Duplicate Score", "Confidence", "Status", "Rule Version", "Explanation", "Top Matching Signals", "Top Conflicting Signals", "Reviewed At", "Review Note", "Merge Operation ID", "Created At", "Updated At"]
OUTREACH_HEADERS = ["Activity ID", "Owner Type", "Owner ID", "Owner Name", "Activity Type", "Channel", "Direction", "Status After", "Contact Value", "Contact Person", "Note", "Follow-up At", "Occurred At", "Created At", "Updated At"]
LOG_HEADERS = ["Log ID", "Collection Job ID", "Job Type", "Job Status", "Log Level", "Message", "Safe Context", "Created At"]


def _csv_project_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    contacts = _group(data["contacts"], "project_id")
    profiles = _group(data["social_profiles"], "project_id")
    campaigns = _group(data["campaigns"], "project_id")
    evidence = _group(data["evidence"], "project_id")
    rels = _group(data["relationships"], "project_id")
    dups = _project_duplicate_map(data["duplicates"])
    rows = []
    for project in data["projects"]:
        project_contacts = contacts.get(project.id, [])
        phones = _contacts_of_type(project_contacts, ["mobile", "phone", "landline"])
        whatsapps = _contacts_of_type(project_contacts, ["whatsapp"])
        emails = _contacts_of_type(project_contacts, ["email"])
        latest_campaign = campaigns.get(project.id, [None])[0]
        rows.append({
            "Project ID": project.id,
            "Project Name": project.name,
            "Developer ID": project.developer_id,
            "Developer Name": project.developer.name if project.developer else "",
            "Developer Classification": project.developer.classification if project.developer else "",
            "Project Classification": project.project_type,
            "Classification Score": "",
            "Review Status": project.review_status,
            "Verification Status": project.verification_status,
            "Project Type": project.project_type,
            "Project Status": project.project_status,
            "Lahore Zone": project.lahore_zone,
            "Address": project.address,
            "Latitude": project.latitude,
            "Longitude": project.longitude,
            "Google Place ID": project.google_place_id,
            "Google Maps URL": project.google_maps_url,
            "Official Website": project.official_website_url,
            "Primary Phone": phones[0].value if phones else "",
            "Other Phones": MULTI_VALUE_SEPARATOR.join(contact.value for contact in phones[1:]),
            "Primary WhatsApp": whatsapps[0].value if whatsapps else "",
            "Other WhatsApp": MULTI_VALUE_SEPARATOR.join(contact.value for contact in whatsapps[1:]),
            "Primary Email": emails[0].value if emails else "",
            "Other Emails": MULTI_VALUE_SEPARATOR.join(contact.value for contact in emails[1:]),
            "Facebook": _profile_url(profiles.get(project.id, []), "facebook"),
            "Instagram": _profile_url(profiles.get(project.id, []), "instagram"),
            "LinkedIn": _profile_url(profiles.get(project.id, []), "linkedin"),
            "X": _profile_url(profiles.get(project.id, []), "x"),
            "YouTube": _profile_url(profiles.get(project.id, []), "youtube"),
            "TikTok": _profile_url(profiles.get(project.id, []), "tiktok"),
            "Campaign Evidence Count": len(campaigns.get(project.id, [])),
            "Latest Campaign Platform": latest_campaign.platform if latest_campaign else "",
            "Latest Campaign CTA": latest_campaign.call_to_action if latest_campaign else "",
            "Latest Campaign Source": latest_campaign.source_url if latest_campaign else "",
            "Pending Relationship": any(rel.status == "candidate" for rel in rels.get(project.id, [])),
            "Pending Duplicate": project.id in dups,
            "Evidence Count": len(evidence.get(project.id, [])),
            "Outreach Status": project.outreach_status,
            "Last Outreach": _iso(project.last_outreach_at),
            "Next Follow-up": _iso(project.next_follow_up_at),
            "Created At": _iso(project.created_at),
            "Updated At": _iso(project.updated_at),
        })
    return rows


def _refined_csv_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [
        {
            "Project ID": record["project_id"],
            "Project Name": record["project_name"],
            "Developer ID": record["developer_id"],
            "Developer Name": record["developer_name"],
            "Developer Phone": record["developer_phone"],
            "Developer Email": record["developer_email"],
            "Developer Website": record["developer_website"],
            "Project Phone": record["project_phone"],
            "Project Email": record["project_email"],
            "Project Website": record["project_website"],
            "Exact Address": record["address"],
            "Latitude": record["latitude"],
            "Longitude": record["longitude"],
            "Google Place ID": record["google_place_id"],
            "Google Maps URL": record["google_maps_url"],
            "Refinement Score": record["refinement_score"],
        }
        for record in data["refined_records"]
    ]


def _developer_contact_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    developers = {developer.id: developer for developer in data["developers"]}
    return [{
        "Contact ID": contact.id,
        "Developer ID": contact.developer_id,
        "Developer Name": developers.get(contact.developer_id).name if developers.get(contact.developer_id) else "",
        "Contact Type": contact.contact_type,
        "Label": contact.label,
        "Original Value": contact.value,
        "Normalized Value": contact.normalized_value,
        "Contact Person": contact.person_name,
        "Designation": contact.designation,
        "Primary": contact.is_primary,
        "Public Business Contact": contact.is_public_business_contact,
        "Verification Status": contact.verification_status,
        "Source URL": contact.source_url,
        "Created At": _iso(contact.created_at),
        "Updated At": _iso(contact.updated_at),
    } for contact in data["contacts"] if contact.developer_id]


def _project_contact_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    projects = {project.id: project for project in data["projects"]}
    return [{
        "Contact ID": contact.id,
        "Project ID": contact.project_id,
        "Project Name": projects.get(contact.project_id).name if projects.get(contact.project_id) else "",
        "Developer Name": projects.get(contact.project_id).developer.name if projects.get(contact.project_id) and projects.get(contact.project_id).developer else "",
        "Contact Type": contact.contact_type,
        "Label": contact.label,
        "Original Value": contact.value,
        "Normalized Value": contact.normalized_value,
        "Contact Person": contact.person_name,
        "Designation": contact.designation,
        "Primary": contact.is_primary,
        "Public Business Contact": contact.is_public_business_contact,
        "Verification Status": contact.verification_status,
        "Source URL": contact.source_url,
        "Created At": _iso(contact.created_at),
        "Updated At": _iso(contact.updated_at),
    } for contact in data["contacts"] if contact.project_id]


def _social_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    developers = {developer.id: developer for developer in data["developers"]}
    projects = {project.id: project for project in data["projects"]}
    return [{
        "Social Profile ID": profile.id,
        "Owner Type": "developer" if profile.developer_id else "project",
        "Owner ID": profile.developer_id or profile.project_id,
        "Owner Name": developers.get(profile.developer_id).name if profile.developer_id and developers.get(profile.developer_id) else projects.get(profile.project_id).name if profile.project_id and projects.get(profile.project_id) else "",
        "Platform": profile.platform,
        "Profile Name": profile.profile_name,
        "Profile URL": profile.profile_url,
        "Normalized URL": profile.normalized_url,
        "Official": profile.is_official,
        "Verification Status": profile.verification_status,
        "Source Evidence Count": 0,
        "Created At": _iso(profile.created_at),
        "Updated At": _iso(profile.updated_at),
    } for profile in data["social_profiles"]]


def _campaign_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    developers = {developer.id: developer for developer in data["developers"]}
    projects = {project.id: project for project in data["projects"]}
    return [{
        "Campaign Evidence ID": item.id,
        "Developer ID": item.developer_id,
        "Developer Name": developers.get(item.developer_id).name if developers.get(item.developer_id) else "",
        "Project ID": item.project_id,
        "Project Name": projects.get(item.project_id).name if projects.get(item.project_id) else "",
        "Platform": item.platform,
        "Campaign Type": item.campaign_type,
        "Advertiser Name": item.advertiser_name,
        "Campaign Text": item.campaign_text,
        "Call to Action": item.call_to_action,
        "Destination URL": item.destination_url,
        "Visible Status": item.visible_status,
        "Verification Status": item.verification_status,
        "First Seen At": _iso(item.first_seen_at),
        "Last Seen At": _iso(item.last_seen_at),
        "Source URL": item.source_url,
        "Social Capture ID": item.social_capture_id,
        "Created At": _iso(item.created_at),
        "Updated At": _iso(item.updated_at),
    } for item in data["campaigns"]]


def _evidence_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    developers = {developer.id: developer for developer in data["developers"]}
    projects = {project.id: project for project in data["projects"]}
    return [{
        "Evidence ID": item.id,
        "Developer ID": item.developer_id,
        "Developer Name": developers.get(item.developer_id).name if developers.get(item.developer_id) else "",
        "Project ID": item.project_id,
        "Project Name": projects.get(item.project_id).name if projects.get(item.project_id) else "",
        "Collection Job ID": item.collection_job_id,
        "Source Type": item.source_type,
        "Source URL": item.source_url,
        "Source Title": item.source_title,
        "Field Name": item.field_name,
        "Extracted Value": item.extracted_value,
        "Relevant Excerpt": item.captured_text,
        "Verification Status": item.verification_status,
        "Collected At": _iso(item.collected_at),
        "Created At": _iso(item.created_at),
    } for item in data["evidence"]]


def _relationship_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    developers = {developer.id: developer for developer in data["developers"]}
    projects = {project.id: project for project in data["projects"]}
    return [{
        "Relationship ID": item.id,
        "Project ID": item.project_id,
        "Project Name": projects.get(item.project_id).name if projects.get(item.project_id) else "",
        "Developer ID": item.developer_id,
        "Developer Name": developers.get(item.developer_id).name if developers.get(item.developer_id) else "",
        "Relationship Type": item.relationship_type,
        "Status": item.status,
        "System Score": item.system_score,
        "Confidence": item.confidence_level,
        "Rule Version": item.rule_version,
        "Explanation": item.explanation,
        "Source URL": item.source_url,
        "Evidence Count": 1 if item.source_evidence_id else 0,
        "Evaluated At": _iso(item.evaluated_at),
        "Reviewed At": _iso(item.reviewed_at),
        "Review Note": item.review_note,
    } for item in data["relationships"]]


def _duplicate_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    developers = {developer.id: developer for developer in data["developers"]}
    projects = {project.id: project for project in data["projects"]}
    rows = []
    for item in data["duplicates"]:
        left_id = item.left_developer_id or item.left_project_id
        right_id = item.right_developer_id or item.right_project_id
        left_name = developers.get(left_id).name if item.entity_type == "developer" and developers.get(left_id) else projects.get(left_id).name if projects.get(left_id) else ""
        right_name = developers.get(right_id).name if item.entity_type == "developer" and developers.get(right_id) else projects.get(right_id).name if projects.get(right_id) else ""
        rows.append({
            "Duplicate Candidate ID": item.id,
            "Entity Type": item.entity_type,
            "Left ID": left_id,
            "Left Name": left_name,
            "Right ID": right_id,
            "Right Name": right_name,
            "Duplicate Score": item.duplicate_score,
            "Confidence": item.confidence_level,
            "Status": item.status,
            "Rule Version": item.rule_version,
            "Explanation": item.explanation,
            "Top Matching Signals": json.dumps(item.signals_json or [], ensure_ascii=False),
            "Top Conflicting Signals": "",
            "Reviewed At": _iso(item.reviewed_at),
            "Review Note": item.review_note,
            "Merge Operation ID": item.merge_operation_id,
            "Created At": _iso(item.created_at),
            "Updated At": _iso(item.updated_at),
        })
    return rows


def _outreach_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    developers = {developer.id: developer for developer in data["developers"]}
    projects = {project.id: project for project in data["projects"]}
    return [{
        "Activity ID": item.id,
        "Owner Type": "developer" if item.developer_id else "project",
        "Owner ID": item.developer_id or item.project_id,
        "Owner Name": developers.get(item.developer_id).name if item.developer_id and developers.get(item.developer_id) else projects.get(item.project_id).name if item.project_id and projects.get(item.project_id) else "",
        "Activity Type": item.activity_type,
        "Channel": item.channel,
        "Direction": item.direction,
        "Status After": item.status_after,
        "Contact Value": item.contact_value,
        "Contact Person": item.contact_person,
        "Note": item.note,
        "Follow-up At": _iso(item.follow_up_at),
        "Occurred At": _iso(item.occurred_at),
        "Created At": _iso(item.created_at),
        "Updated At": _iso(item.updated_at),
    } for item in data["outreach"]]


def _collection_log_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    return [{
        "Log ID": item.id,
        "Collection Job ID": item.collection_job_id,
        "Job Type": "",
        "Job Status": "",
        "Log Level": item.level,
        "Message": item.message,
        "Safe Context": json.dumps(item.context_json or {}, ensure_ascii=False),
        "Created At": _iso(item.created_at),
    } for item in data["collection_logs"]]


def _review_queue(projects: list[Project], developers: list[Developer], relationships: list[ProjectDeveloperRelationship], duplicates: list[DuplicateCandidate], captures: list[SocialCapture]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for project in projects:
        if project.review_status in {"unreviewed", "needs_review"} or not project.developer_id:
            rows.append({"Entity Type": "project", "Entity ID": project.id, "Entity Name": project.name, "Review Status": project.review_status, "Effective Classification": project.project_type, "Suggested Classification": "", "Classification Score": "", "Pending Relationship": False, "Pending Duplicate": False, "Missing Developer": not project.developer_id, "Missing Phone": "", "Missing WhatsApp": "", "Missing Email": "", "Missing Website": not bool(project.official_website_url), "Missing Coordinates": not bool(project.latitude and project.longitude), "Conflicting Evidence": False, "Latest Warning": "", "Last Reviewed At": _iso(project.last_reviewed_at), "Created At": _iso(project.created_at), "Updated At": _iso(project.updated_at)})
    for developer in developers:
        if developer.review_status in {"unreviewed", "needs_review"}:
            rows.append({"Entity Type": "developer", "Entity ID": developer.id, "Entity Name": developer.name, "Review Status": developer.review_status, "Effective Classification": developer.classification, "Suggested Classification": "", "Classification Score": "", "Pending Relationship": False, "Pending Duplicate": False, "Missing Developer": False, "Missing Phone": "", "Missing WhatsApp": "", "Missing Email": "", "Missing Website": not bool(developer.website_url), "Missing Coordinates": False, "Conflicting Evidence": False, "Latest Warning": "", "Last Reviewed At": _iso(developer.last_reviewed_at), "Created At": _iso(developer.created_at), "Updated At": _iso(developer.updated_at)})
    for rel in relationships:
        if rel.status == "candidate":
            rows.append({"Entity Type": "relationship", "Entity ID": rel.id, "Entity Name": f"Project {rel.project_id} -> Developer {rel.developer_id}", "Review Status": rel.status, "Effective Classification": rel.relationship_type, "Suggested Classification": "", "Classification Score": rel.system_score, "Pending Relationship": True, "Pending Duplicate": False, "Missing Developer": False, "Missing Phone": False, "Missing WhatsApp": False, "Missing Email": False, "Missing Website": False, "Missing Coordinates": False, "Conflicting Evidence": False, "Latest Warning": rel.explanation, "Last Reviewed At": _iso(rel.reviewed_at), "Created At": _iso(rel.created_at), "Updated At": _iso(rel.updated_at)})
    for dup in duplicates:
        if dup.status == "pending":
            rows.append({"Entity Type": "duplicate_candidate", "Entity ID": dup.id, "Entity Name": dup.entity_type, "Review Status": dup.status, "Effective Classification": "", "Suggested Classification": "", "Classification Score": dup.duplicate_score, "Pending Relationship": False, "Pending Duplicate": True, "Missing Developer": False, "Missing Phone": False, "Missing WhatsApp": False, "Missing Email": False, "Missing Website": False, "Missing Coordinates": False, "Conflicting Evidence": False, "Latest Warning": dup.explanation, "Last Reviewed At": _iso(dup.reviewed_at), "Created At": _iso(dup.created_at), "Updated At": _iso(dup.updated_at)})
    for capture in captures:
        if capture.review_status == "unassigned":
            rows.append({"Entity Type": "social_capture", "Entity ID": capture.id, "Entity Name": capture.profile_name or capture.page_title or capture.source_url, "Review Status": capture.review_status, "Effective Classification": capture.platform, "Suggested Classification": "", "Classification Score": "", "Pending Relationship": False, "Pending Duplicate": False, "Missing Developer": not capture.developer_id, "Missing Phone": False, "Missing WhatsApp": False, "Missing Email": False, "Missing Website": False, "Missing Coordinates": False, "Conflicting Evidence": False, "Latest Warning": "", "Last Reviewed At": _iso(capture.reviewed_at), "Created At": _iso(capture.created_at), "Updated At": _iso(capture.updated_at)})
    return rows


def _excluded_rows(data: dict[str, Any]) -> list[dict[str, Any]]:
    rows = []
    for project in data["projects"]:
        if project.review_status in {"rejected", "excluded"}:
            rows.append({"Entity Type": "project", "Entity ID": project.id, "Entity Name": project.name, "Classification": project.project_type, "Review Status": project.review_status, "Exclusion or Rejection Reason": project.review_note, "Classification Score": "", "Top Positive Signal": "", "Top Negative Signal": "", "Website": project.official_website_url, "Phone": "", "Lahore Zone": project.lahore_zone, "Evidence Count": 0, "Reviewed At": _iso(project.last_reviewed_at), "Updated At": _iso(project.updated_at)})
    for developer in data["developers"]:
        if developer.review_status in {"rejected", "excluded"} or developer.classification in {"broker", "marketing_partner"}:
            rows.append({"Entity Type": "developer", "Entity ID": developer.id, "Entity Name": developer.name, "Classification": developer.classification, "Review Status": developer.review_status, "Exclusion or Rejection Reason": developer.review_note, "Classification Score": "", "Top Positive Signal": "", "Top Negative Signal": "", "Website": developer.website_url, "Phone": "", "Lahore Zone": "", "Evidence Count": 0, "Reviewed At": _iso(developer.last_reviewed_at), "Updated At": _iso(developer.updated_at)})
    return rows


def _dashboard_rows(artifact: ExportArtifact, data: dict[str, Any], stats: ExportStats) -> list[dict[str, Any]]:
    summary = data["summary"]
    return [
        {"Metric": "Export ID", "Value": artifact.id},
        {"Metric": "Format", "Value": artifact.format},
        {"Metric": "Scope", "Value": artifact.scope},
        {"Metric": "Schema Version", "Value": EXPORT_SCHEMA_VERSION},
        {"Metric": "Projects Exported", "Value": summary["projects_exported"]},
        {"Metric": "Developers Exported", "Value": summary["developers_exported"]},
        {"Metric": "Contacts Exported", "Value": summary["contacts_exported"]},
        {"Metric": "Social Profiles Exported", "Value": summary["social_profiles_exported"]},
        {"Metric": "Campaign Evidence Exported", "Value": summary["campaign_evidence_exported"]},
        {"Metric": "Source Evidence Exported", "Value": summary["source_evidence_exported"]},
        {"Metric": "Formula Protected Cells", "Value": stats.protected_cells},
        {"Metric": "Truncated Cells", "Value": stats.truncated_cells},
    ]


def _metadata_rows(artifact: ExportArtifact, data: dict[str, Any], stats: ExportStats) -> list[dict[str, Any]]:
    return [
        {"Key": "Export ID", "Value": artifact.id},
        {"Key": "Collection Job ID", "Value": artifact.collection_job_id},
        {"Key": "Format", "Value": artifact.format},
        {"Key": "Scope", "Value": artifact.scope},
        {"Key": "Requested At", "Value": (artifact.filter_snapshot_json or {}).get("requested_at")},
        {"Key": "Generated At", "Value": _iso(utc_now())},
        {"Key": "Application Version", "Value": "0.9.0"},
        {"Key": "Export Schema Version", "Value": EXPORT_SCHEMA_VERSION},
        {"Key": "Project Filter Snapshot", "Value": json.dumps((artifact.filter_snapshot_json or {}).get("project_filters", {}), ensure_ascii=False)},
        {"Key": "Developer Filter Snapshot", "Value": json.dumps((artifact.filter_snapshot_json or {}).get("developer_filters", {}), ensure_ascii=False)},
        {"Key": "Export Options", "Value": json.dumps(artifact.options_json or {}, ensure_ascii=False)},
        {"Key": "Primary Row Count", "Value": data["primary_rows"]},
        {"Key": "Total Rows", "Value": data["total_rows"]},
        {"Key": "Truncated Cell Count", "Value": stats.truncated_cells},
        {"Key": "Protected Formula Cell Count", "Value": stats.protected_cells},
        {"Key": "Warnings", "Value": json.dumps(stats.warnings or [], ensure_ascii=False)},
    ]


def _json_payload(artifact: ExportArtifact, data: dict[str, Any]) -> dict[str, Any]:
    contacts_by_dev = _group(data["contacts"], "developer_id")
    contacts_by_project = _group(data["contacts"], "project_id")
    profiles_by_dev = _group(data["social_profiles"], "developer_id")
    profiles_by_project = _group(data["social_profiles"], "project_id")
    projects_by_dev = _group(data["projects"], "developer_id")
    campaigns_by_project = _group(data["campaigns"], "project_id")
    rels_by_project = _group(data["relationships"], "project_id")
    dup_projects = _project_duplicate_map(data["duplicates"])
    evidence_by_dev = _group(data["evidence"], "developer_id")
    evidence_by_project = _group(data["evidence"], "project_id")
    outreach_by_dev = _group(data["outreach"], "developer_id")
    outreach_by_project = _group(data["outreach"], "project_id")
    seen_projects: set[int] = set()
    developers = []
    for developer in data["developers"]:
        nested_projects = []
        for project in projects_by_dev.get(developer.id, []):
            nested_projects.append(_json_project(project, contacts_by_project, profiles_by_project, campaigns_by_project, rels_by_project, dup_projects, outreach_by_project, evidence_by_project, artifact))
            seen_projects.add(project.id)
        developers.append({
            "id": developer.id,
            "name": developer.name,
            "legal_name": developer.legal_name,
            "classification": {"effective": developer.classification, "suggested": developer.classification, "score": None, "confidence": None, "review_status": None, "rule_version": None},
            "review": {"status": developer.review_status, "note": developer.review_note, "last_reviewed_at": _iso(developer.last_reviewed_at)},
            "website_url": developer.website_url,
            "office_address": developer.office_address,
            "contacts": [_json_contact(contact) for contact in contacts_by_dev.get(developer.id, [])],
            "social_profiles": [_json_profile(profile) for profile in profiles_by_dev.get(developer.id, [])],
            "projects": nested_projects,
            "outreach": {"status": developer.outreach_status, "last_outreach_at": _iso(developer.last_outreach_at), "next_follow_up_at": _iso(developer.next_follow_up_at), "activities": [_json_outreach(item) for item in outreach_by_dev.get(developer.id, [])]},
            "source_evidence": [_json_evidence(item) for item in evidence_by_dev.get(developer.id, [])] if (artifact.options_json or {}).get("include_source_evidence") else {"source_evidence_count": len(evidence_by_dev.get(developer.id, []))},
        })
    unassigned = [_json_project(project, contacts_by_project, profiles_by_project, campaigns_by_project, rels_by_project, dup_projects, outreach_by_project, evidence_by_project, artifact) for project in data["projects"] if project.id not in seen_projects]
    return {
        "schema_version": EXPORT_SCHEMA_VERSION,
        "metadata": {"export_id": artifact.id, "scope": artifact.scope, "requested_at": (artifact.filter_snapshot_json or {}).get("requested_at"), "generated_at": _iso(utc_now()), "application": "Alduor Lahore Project Discovery Agent", "application_version": "0.9.0", "schema_version": EXPORT_SCHEMA_VERSION, "intelligence_rule_version": "m6-v1", "timezone": "UTC", "options": artifact.options_json or {}},
        "filters": artifact.filter_snapshot_json or {},
        "summary": data["summary"],
        "developers": developers,
        "unassigned_projects": unassigned,
        "review_queue": data["review_queue"],
        "duplicate_candidates": _duplicate_rows(data),
        "social_captures": [_json_social_capture(item) for item in data["social_captures"]],
        "collection_jobs": [],
    }


def _refined_json_payload(artifact: ExportArtifact, data: dict[str, Any]) -> dict[str, Any]:
    projects = []
    for record in data["refined_records"]:
        projects.append(
            {
                "project_id": record["project_id"],
                "project_name": record["project_name"],
                "developer": {
                    "developer_id": record["developer_id"],
                    "name": record["developer_name"],
                    "phone": record["developer_phone"],
                    "email": record["developer_email"],
                    "website": record["developer_website"],
                },
                "project_contact": {
                    "phone": record["project_phone"],
                    "email": record["project_email"],
                    "website": record["project_website"],
                },
                "location": {
                    "address": record["address"],
                    "latitude": record["latitude"],
                    "longitude": record["longitude"],
                    "google_place_id": record["google_place_id"],
                    "google_maps_url": record["google_maps_url"],
                },
                "quality": {
                    "refinement_score": record["refinement_score"],
                    "required_fields_complete": True,
                },
            }
        )
    return {
        "schema_version": REFINEMENT_SCHEMA_VERSION,
        "generated_at": _iso(utc_now()),
        "scope": artifact.scope,
        "summary": {
            "raw_projects": data["summary"].get("raw_projects", 0),
            "likely_real_estate_projects": data["summary"].get("likely_real_estate_projects", 0),
            "export_ready_projects": len(projects),
            "excluded_incomplete_projects": data["summary"].get("excluded_incomplete_projects", 0),
        },
        "projects": projects,
    }


def _json_project(project: Project, contacts: dict[int, list[Contact]], profiles: dict[int, list[SocialProfile]], campaigns: dict[int, list[CampaignEvidence]], relationships: dict[int, list[ProjectDeveloperRelationship]], duplicates: dict[int, list[DuplicateCandidate]], outreach: dict[int, list[OutreachActivity]], evidence: dict[int, list[SourceEvidence]], artifact: ExportArtifact) -> dict[str, Any]:
    return {
        "id": project.id,
        "name": project.name,
        "developer_id": project.developer_id,
        "project_type": project.project_type,
        "project_status": project.project_status,
        "classification": {"effective": project.project_type, "suggested": project.project_type, "score": None, "confidence": None},
        "review": {"status": project.review_status, "note": project.review_note},
        "location": {"lahore_zone": project.lahore_zone, "address": project.address, "latitude": project.latitude, "longitude": project.longitude},
        "google": {"place_id": project.google_place_id, "maps_url": project.google_maps_url},
        "official_website_url": project.official_website_url,
        "contacts": [_json_contact(contact) for contact in contacts.get(project.id, [])],
        "social_profiles": [_json_profile(profile) for profile in profiles.get(project.id, [])],
        "campaign_evidence": [_json_campaign(item) for item in campaigns.get(project.id, [])],
        "relationships": _relationship_rows({"relationships": relationships.get(project.id, []), "developers": [], "projects": [project]}),
        "duplicate_candidates": [item.id for item in duplicates.get(project.id, [])],
        "outreach": {"status": project.outreach_status, "last_outreach_at": _iso(project.last_outreach_at), "next_follow_up_at": _iso(project.next_follow_up_at), "activities": [_json_outreach(item) for item in outreach.get(project.id, [])]},
        "source_evidence": [_json_evidence(item) for item in evidence.get(project.id, [])] if (artifact.options_json or {}).get("include_source_evidence") else {"source_evidence_count": len(evidence.get(project.id, []))},
    }


def _json_contact(contact: Contact) -> dict[str, Any]:
    return {"id": contact.id, "contact_type": contact.contact_type, "label": contact.label, "value": contact.value, "normalized_value": contact.normalized_value, "verification_status": contact.verification_status, "source_url": contact.source_url}


def _json_profile(profile: SocialProfile) -> dict[str, Any]:
    return {"id": profile.id, "platform": profile.platform, "profile_name": profile.profile_name, "profile_url": profile.profile_url, "normalized_url": profile.normalized_url, "is_official": profile.is_official, "verification_status": profile.verification_status}


def _json_campaign(item: CampaignEvidence) -> dict[str, Any]:
    return {"id": item.id, "platform": item.platform, "campaign_type": item.campaign_type, "advertiser_name": item.advertiser_name, "campaign_text": item.campaign_text, "call_to_action": item.call_to_action, "destination_url": item.destination_url, "visible_status": item.visible_status, "verification_status": item.verification_status, "first_seen_at": _iso(item.first_seen_at), "last_seen_at": _iso(item.last_seen_at), "source_url": item.source_url, "social_capture_id": item.social_capture_id}


def _json_evidence(item: SourceEvidence) -> dict[str, Any]:
    return {"id": item.id, "source_type": item.source_type, "source_url": item.source_url, "source_title": item.source_title, "field_name": item.field_name, "extracted_value": item.extracted_value, "relevant_excerpt": item.captured_text, "verification_status": item.verification_status, "collected_at": _iso(item.collected_at)}


def _json_outreach(item: OutreachActivity) -> dict[str, Any]:
    return {"id": item.id, "activity_type": item.activity_type, "channel": item.channel, "direction": item.direction, "status_after": item.status_after, "contact_value": item.contact_value, "contact_person": item.contact_person, "note": item.note, "follow_up_at": _iso(item.follow_up_at), "occurred_at": _iso(item.occurred_at)}


def _json_social_capture(item: SocialCapture) -> dict[str, Any]:
    return {"id": item.id, "platform": item.platform, "page_kind": item.page_kind, "source_url": item.source_url, "canonical_url": item.canonical_url, "profile_name": item.profile_name, "username": item.username, "review_status": item.review_status, "captured_at": _iso(item.captured_at)}


def _group(items: list[Any], attr: str) -> dict[int, list[Any]]:
    grouped: dict[int, list[Any]] = {}
    for item in items:
        key = getattr(item, attr, None)
        if key is not None:
            grouped.setdefault(key, []).append(item)
    return grouped


def _primary_contact(contacts: list[Contact], types: list[str]) -> Contact | None:
    filtered = _contacts_of_type(contacts, types)
    return filtered[0] if filtered else None


def _contacts_of_type(contacts: list[Contact], types: list[str]) -> list[Contact]:
    priority = {value: index for index, value in enumerate(types)}
    return sorted([contact for contact in contacts if contact.contact_type in types], key=lambda item: (not item.is_primary, item.verification_status != "verified", priority.get(item.contact_type, 99), item.id))


def _profile_url(profiles: list[SocialProfile], platform: str) -> str:
    matches = [profile for profile in profiles if profile.platform == platform]
    matches.sort(key=lambda item: (not item.is_official, item.verification_status != "verified", item.id))
    return matches[0].profile_url if matches else ""


def _project_duplicate_map(duplicates: list[DuplicateCandidate]) -> dict[int, list[DuplicateCandidate]]:
    mapped: dict[int, list[DuplicateCandidate]] = {}
    for item in duplicates:
        for project_id in (item.left_project_id, item.right_project_id):
            if project_id:
                mapped.setdefault(project_id, []).append(item)
    return mapped


def _developer_duplicate_map(duplicates: list[DuplicateCandidate]) -> dict[int, list[DuplicateCandidate]]:
    mapped: dict[int, list[DuplicateCandidate]] = {}
    for item in duplicates:
        for developer_id in (item.left_developer_id, item.right_developer_id):
            if developer_id:
                mapped.setdefault(developer_id, []).append(item)
    return mapped


def _progress(context: Any | None, db: Session, artifact: ExportArtifact, phase: str, message: str, *, total: int | None = None, processed: int | None = None, created: int | None = None) -> None:
    _check_cancelled(context)
    if context is not None:
        db.commit()
        kwargs: dict[str, Any] = {"phase": phase, "message": message}
        if total is not None:
            kwargs["total_items"] = total
        if processed is not None:
            kwargs["processed_delta"] = processed
        if created is not None:
            kwargs["created_delta"] = created
        context.update_progress(**kwargs)
        db.refresh(artifact)
    artifact.updated_at = utc_now()
    db.flush()


def _check_cancelled(context: Any | None) -> None:
    if context is not None:
        context.check_cancelled()


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _is_formula_like(value: str) -> bool:
    stripped = value.lstrip(" \t\r\n")
    return bool(stripped) and stripped[0] in FORMULA_PREFIXES


def _iso(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        return value
    if isinstance(value, datetime):
        return _as_utc(value).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    return str(value)


def _as_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value.astimezone(UTC)


def _truthy(value: Any) -> bool:
    return value is True or str(value).lower() in {"true", "1", "yes"}
